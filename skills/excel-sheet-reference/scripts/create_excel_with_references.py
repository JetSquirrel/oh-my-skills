#!/usr/bin/env python3
"""
Excel Sheet Reference Example Script

This script demonstrates how to create Excel files with multiple sheets
and cross-sheet formula references using openpyxl.

Usage:
    python create_excel_with_references.py

Requirements:
    pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_basic_example():
    """
    Create a basic example matching the issue description.
    Sheet1: Raw fruit data
    Sheet2: Summary with COUNTIFS formulas
    """
    print("Creating basic fruit summary example...")
    
    wb = Workbook()
    
    # Sheet 1: Source Data
    sheet1 = wb.active
    sheet1.title = "sheet1"
    
    # Add header with styling
    sheet1['A1'] = 'fruit'
    sheet1['A1'].font = Font(bold=True)
    sheet1['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add fruit data
    sheet1['A2'] = 'apple'
    sheet1['A3'] = 'apple'
    sheet1['A4'] = 'banana'
    
    # Sheet 2: Summary with formulas
    sheet2 = wb.create_sheet("sheet2")
    
    # Add headers with styling
    sheet2['A1'] = 'fruit'
    sheet2['B1'] = 'count'
    
    for cell in ['A1', 'B1']:
        sheet2[cell].font = Font(bold=True)
        sheet2[cell].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add summary data with COUNTIFS formulas
    sheet2['A2'] = 'apple'
    sheet2['B2'] = '=COUNTIFS(sheet1!A:A,A2)'
    
    sheet2['A3'] = 'banana'
    sheet2['B3'] = '=COUNTIFS(sheet1!A:A,A3)'
    
    # Adjust column widths
    sheet1.column_dimensions['A'].width = 15
    sheet2.column_dimensions['A'].width = 15
    sheet2.column_dimensions['B'].width = 10
    
    wb.save('basic_fruit_summary.xlsx')
    print("✓ Created: basic_fruit_summary.xlsx")


def create_advanced_example():
    """
    Create an advanced example with VLOOKUP and multiple formulas.
    """
    print("\nCreating advanced example with VLOOKUP...")
    
    wb = Workbook()
    
    # Sheet 1: Product Database
    products = wb.active
    products.title = "Products"
    
    # Headers
    headers = ['Product', 'Price', 'Category', 'Stock']
    for idx, header in enumerate(headers, start=1):
        cell = products.cell(row=1, column=idx)
        cell.value = header
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Sample data
    product_data = [
        ['apple', 1.50, 'fruit', 100],
        ['banana', 0.80, 'fruit', 150],
        ['carrot', 1.20, 'vegetable', 80],
        ['orange', 2.00, 'fruit', 120],
        ['broccoli', 2.50, 'vegetable', 60],
    ]
    
    for row_idx, row_data in enumerate(product_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            products.cell(row=row_idx, column=col_idx, value=value)
    
    # Sheet 2: Sales Orders with VLOOKUP
    orders = wb.create_sheet("Orders")
    
    # Headers
    order_headers = ['Product', 'Quantity', 'Unit Price', 'Total', 'Category']
    for idx, header in enumerate(order_headers, start=1):
        cell = orders.cell(row=1, column=idx)
        cell.value = header
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Order data
    order_items = ['apple', 'banana', 'carrot', 'orange']
    quantities = [5, 10, 3, 8]
    
    for row_idx, (product, qty) in enumerate(zip(order_items, quantities), start=2):
        orders[f'A{row_idx}'] = product
        orders[f'B{row_idx}'] = qty
        # VLOOKUP to get price from Products sheet
        orders[f'C{row_idx}'] = f'=VLOOKUP(A{row_idx},Products!A:D,2,FALSE)'
        # Calculate total
        orders[f'D{row_idx}'] = f'=B{row_idx}*C{row_idx}'
        # VLOOKUP to get category
        orders[f'E{row_idx}'] = f'=VLOOKUP(A{row_idx},Products!A:D,3,FALSE)'
    
    # Add totals row
    total_row = len(order_items) + 2
    orders[f'A{total_row}'] = 'TOTAL'
    orders[f'A{total_row}'].font = Font(bold=True)
    orders[f'D{total_row}'] = f'=SUM(D2:D{total_row-1})'
    orders[f'D{total_row}'].font = Font(bold=True)
    
    # Sheet 3: Category Summary
    summary = wb.create_sheet("CategorySummary")
    
    # Headers
    summary['A1'] = 'Category'
    summary['B1'] = 'Product Count'
    summary['C1'] = 'Avg Price'
    
    for cell in ['A1', 'B1', 'C1']:
        summary[cell].font = Font(bold=True)
        summary[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        summary[cell].alignment = Alignment(horizontal='center')
    
    # Summary data with formulas
    categories = ['fruit', 'vegetable']
    
    for row_idx, category in enumerate(categories, start=2):
        summary[f'A{row_idx}'] = category
        # Count products in category
        summary[f'B{row_idx}'] = f'=COUNTIFS(Products!C:C,A{row_idx})'
        # Average price for category
        summary[f'C{row_idx}'] = f'=AVERAGEIF(Products!C:C,A{row_idx},Products!B:B)'
    
    # Adjust column widths for all sheets
    for sheet in [products, orders, summary]:
        for column in sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save('advanced_orders_with_vlookup.xlsx')
    print("✓ Created: advanced_orders_with_vlookup.xlsx")


def create_match_index_example():
    """
    Create an example using MATCH and INDEX functions.
    """
    print("\nCreating MATCH and INDEX example...")
    
    wb = Workbook()
    
    # Data sheet
    data = wb.active
    data.title = "EmployeeData"
    
    # Headers
    headers = ['Name', 'Age', 'Department', 'Salary']
    for idx, header in enumerate(headers, start=1):
        cell = data.cell(row=1, column=idx)
        cell.value = header
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Employee data
    employees = [
        ['Alice', 25, 'Engineering', 75000],
        ['Bob', 30, 'Marketing', 65000],
        ['Charlie', 35, 'Engineering', 85000],
        ['Diana', 28, 'HR', 60000],
        ['Eve', 32, 'Marketing', 70000],
    ]
    
    for row_idx, emp in enumerate(employees, start=2):
        for col_idx, value in enumerate(emp, start=1):
            data.cell(row=row_idx, column=col_idx, value=value)
    
    # Lookup sheet
    lookup = wb.create_sheet("Lookup")
    
    # Headers
    lookup['A1'] = 'Search Name'
    lookup['B1'] = 'Position'
    lookup['C1'] = 'Age'
    lookup['D1'] = 'Department'
    lookup['E1'] = 'Salary'
    
    for cell_ref in ['A1', 'B1', 'C1', 'D1', 'E1']:
        lookup[cell_ref].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        lookup[cell_ref].font = Font(bold=True, color="FFFFFF")
    
    # Example lookups
    search_names = ['Bob', 'Diana', 'Alice']
    
    for row_idx, name in enumerate(search_names, start=2):
        lookup[f'A{row_idx}'] = name
        # MATCH to find row position
        lookup[f'B{row_idx}'] = f'=MATCH(A{row_idx},EmployeeData!A:A,0)'
        # INDEX-MATCH to get age
        lookup[f'C{row_idx}'] = f'=INDEX(EmployeeData!B:B,MATCH(A{row_idx},EmployeeData!A:A,0))'
        # INDEX-MATCH to get department
        lookup[f'D{row_idx}'] = f'=INDEX(EmployeeData!C:C,MATCH(A{row_idx},EmployeeData!A:A,0))'
        # INDEX-MATCH to get salary
        lookup[f'E{row_idx}'] = f'=INDEX(EmployeeData!D:D,MATCH(A{row_idx},EmployeeData!A:A,0))'
    
    # Adjust column widths
    for sheet in [data, lookup]:
        for col in range(1, 6):
            sheet.column_dimensions[get_column_letter(col)].width = 15
    
    wb.save('employee_lookup_with_match.xlsx')
    print("✓ Created: employee_lookup_with_match.xlsx")


def main():
    """Main function to create all examples."""
    print("=" * 60)
    print("Excel Sheet Reference Examples")
    print("=" * 60)
    
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is not installed.")
        print("Please install it with: pip install openpyxl")
        return
    
    # Create all examples
    create_basic_example()
    create_advanced_example()
    create_match_index_example()
    
    print("\n" + "=" * 60)
    print("All examples created successfully!")
    print("=" * 60)
    print("\nCreated files:")
    print("  1. basic_fruit_summary.xlsx")
    print("  2. advanced_orders_with_vlookup.xlsx")
    print("  3. employee_lookup_with_match.xlsx")
    print("\nOpen these files in Excel to see the formulas in action.")


if __name__ == "__main__":
    main()
